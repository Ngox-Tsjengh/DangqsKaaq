from openpyxl import load_workbook          #讀入xlsx文件
from openpyxl.utils import range_boundaries
import json                                 #
# import re                                   #字符替換

workbook = load_workbook(filename="KuangxYonh.xlsx")
sheet = workbook["KuangxYonh"]

# df = DataFrame(workbook.values)
data = []
phengs = []

for row in sheet.iter_rows(min_row=3, max_row=8546,
                           min_col=0, max_col=13,
                           values_only=True):
    datum = {
        "轄字": row[11],
        "聲符": row[0],
        "上古擬音": row[4],
        "聲母": row[5],
        "音節類型": row[2],
        "等": row[6],
        "開合": row[7],
        "韻母": row[8],
        "聲調": row[9],
    }
    data.append(datum)

    qim = str(row[4])
    qim = qim.replace('(', '')
    qim = qim.replace(')', '')
    qim = qim.replace('[', '')
    qim = qim.replace(']', '')

    qim = qim.replace('ʰ', 'h')
    qim = qim.replace('ɹ', 'r')
    qim = qim.replace('w̥', 'hw')
    qim = qim.replace('l̥', 'hl')
    qim = qim.replace('m̥', 'hm')
    qim = qim.replace('n̥', 'hn')
    qim = qim.replace('r̥', 'hr')
    qim = qim.replace('ŋ̊', 'hy')
    qim = qim.replace('ŋ', 'y')
    qim = qim.replace('ʔ', 'q')
    qim = qim.replace('ɡ', 'g')

    qim = qim.replace('a̠', 'aa')
    qim = qim.replace('e̠', 'ee')
    qim = qim.replace('i̠', 'ii')
    qim = qim.replace('o̠', 'oo')
    qim = qim.replace('u̠', 'uu')
    qim = qim.replace('ə̠', 'əə')
    qim = qim.replace('ə', 'v')

    qim = qim.replace('ɬ', 'tl')
    qim = qim.replace('ɫ', 'l')
    qim = qim.replace('ɫ', 'l')
    qim = qim.replace('ˁ', 'q')

    for char in row[11]:
        if (char == "<") or (char == ">"):
            continue
        if (qim == "None"):
            continue
        pheng = {str(char): qim}
        phengs.append(pheng)

with open('data.json', 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=4)

with open('phengs.json', 'w', encoding='utf-8') as f:
    json.dump(phengs, f, ensure_ascii=False, indent=4)
    
dict_head = """---
name: dayqskaaq
version: \"2022.10.15\"
sort: by_weight
use_preset_vocabulary: true
...

"""

with open('dayqskaaq.dict.yaml', 'w') as f:
    f.writelines(dict_head)
    for pheng in phengs:
        string = [ f'{key}\t{pheng[key]}' for key in pheng ]
        for st in string:
            f.write(f'{st}\n')

# with open('dangqskaaq.dict.yaml', 'w') as f:
    # list_of_strings = [ f'{key}\t{qims[key]}' for key in qims ]
    # for st in list_of_strings:
        # f.write(f'{st}\n')

